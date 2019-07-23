import openpyxl
from openpyxl import Workbook
import time
from datetime import date 

#path = 'D:\Projects\python_practice\python-excel\Ritik_Progress.xlsx'
sheet = openpyxl.load_workbook('Ritik_Progress.xlsx')
ws = sheet.active

print('Hey Ritik Track your progress at robotronix for today!')

time_delta = input('Enter time duration : ')
progress = input('Enter today\'s progress : ')
date = date.today().strftime('%d/%m/%Y')
print('Your response have been saved. Have a great day!')

for i in range(2,500):
	if ws.cell(row=i, column=1).value == None:
		print('I am inside the loop')
		ws.cell(row=i,column=1).value = date 
		break

for i in range(2,500):
	if ws.cell(row=i, column=3).value == None:
		ws.cell(row=i,column=3).value = time_delta 
		break

for i in range(2,500):
	if ws.cell(row=i, column=5).value == None:
		ws.cell(row=i,column=5).value = progress
		break

sheet.save('Ritik_Progress.xlsx')
